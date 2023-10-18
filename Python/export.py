import openpyxl
from openpyxl.styles import PatternFill
import csv

import manoutils


def createExcelWorkBook(name, startAscending, startTransverse, startDescending, startSigmoid, startRectum, endRectum,
                        data, commentsDict, distance, first_sensor, last_sensor):
    header = [
        ['Time', 'Ant/Retr', 'Amplitude', 'Velocity mm/s', 'startSensor', 'endSensor', 'lengthContraction']
    ]
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    headerSize = len(header[0])
    # add regions
    count = 0
    regionsline = [None, None, None, None, None, None, None, 'Ascending']
    while startAscending + count < startTransverse:
        if count != 0:
            regionsline.append(None)
        header[0].append(startAscending + count)
        count += 1
    count = 0
    regionsline.append('Transverse')
    while startTransverse + count < startDescending:
        if count != 0:
            regionsline.append(None)
        header[0].append(startTransverse + count)
        count += 1
    count = 0
    regionsline.append('Descending')
    while startDescending + count < startSigmoid:
        if count != 0:
            regionsline.append(None)
        header[0].append(startDescending + count)
        count += 1
    count = 0
    regionsline.append('Sigmoid')
    while startSigmoid + count < startRectum:
        if count != 0:
            regionsline.append(None)
        header[0].append(startSigmoid + count)
        count += 1
    count = 0
    regionsline.append('Rectum')
    while not startRectum + count > endRectum:
        if count != 0:
            regionsline.append(None)
        header[0].append(startRectum + count)
        count += 1

    # blijkbaar kan excel enkel lists of list aannemen, header =[regionsline, header] werkt ni for some reason
    regionsline = [regionsline]

    # try to adapt data to header format
    contractions = []
    contractionsDict = dict()
    for line in data:
        # calculate max amplitude
        max_y_values = []
        contractionsDict = dict()
        # to see position where y max is
        xvaluesDict = dict()
        count = 0
        for entry in line['sequences']:
            max_y_values.append(max(entry, key=lambda x: x[1])[1])
            # make a dict with amp per sensor
            for point in entry:
                count += 1
                try:
                    if contractionsDict[point[0]] < point[1]:
                        contractionsDict[point[0]] = point[1]
                        xvaluesDict[point[0]] = count
                except:
                    contractionsDict[point[0]] = point[1]
                    xvaluesDict[point[0]] = count
        maxAmp = max(max_y_values)
        # calculate velocity
        gran = manoutils.get_granularity_factor()
        time = gran * len(line['sequences'])
        # last sensor - first * distance between each sensor = total distance, time is in decaseconds => *10, distance in mm => *10
        velocity = (line['sequences'][-1][-1][0] - line['sequences'][0][0][0]) * distance / time * 100
        # decide ant/retr
        if (velocity > 0):
            direction = 'Ant'
        else:
            direction = 'Retr'
        startSensor = line['sequences'][0][0][0]
        endSensor = line['sequences'][-1][-1][0]

        # simultaan = gehele colon en snelheid infinite => all highpoints zelfde measurement
        if startSensor == first_sensor and endSensor == last_sensor:
            # check hoogte punt van elke entry op zelfde meting sinds start
            start = 0
            count = 0
            for entry in xvaluesDict:
                count += 1
                if start == 0:
                    start = xvaluesDict[entry]
                elif xvaluesDict[entry] != start:
                    break
                else:
                    continue
            if count - 1 == last_sensor:
                direction = 'Simul'

        contract = [manoutils.convertTimeToText(manoutils.get_granularity_factor() * line['measure_number']), direction,
                    maxAmp, velocity, startSensor, endSensor, line['length']]
        count = startAscending - 1
        # add none/amplitude per sensor
        while count < endRectum:
            try:
                contract.append(contractionsDict[count])
            except:
                contract.append(None)
            count += 1
        contractions.append(contract)
    # also export to csv
    exportToCsv(contractions, name)
    # try to add comments to data
    try:
        # converts commentsDict to list, adds them to the data, then sorts it by time
        commentsList = [[manoutils.convertTimeToText(key), value] for key, value in commentsDict.items()]
        contractions += commentsList
        contractions = sorted(contractions, key=lambda x: x[0])
    except:
        print("error in comments toevoegen")
    # try to append data to the header
    try:
        for line in contractions:
            header.append(line)
    except:
        print("error in datalijn toevoegen")
    # Write all data to xlsx
    for row in regionsline:
        worksheet.append(row)
    for row in header:
        worksheet.append(row)
    # length of regions
    lenAsc = startTransverse - startAscending
    lenTrans = startDescending - startTransverse
    lenDesc = startSigmoid - startDescending
    lenSig = startRectum - startSigmoid
    lenRect = endRectum - startRectum + 1
    # decide where regions stop
    stopAsc = headerSize + lenAsc
    stopTrans = stopAsc + lenTrans
    stopDesc = stopTrans + lenDesc
    stopSig = stopDesc + lenSig
    stopRect = stopSig + lenRect

    # colour the regions to distinguish them
    fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    for row in worksheet.iter_rows(min_row=1, max_row=1, min_col=headerSize + 1, max_col=stopAsc):
        for cell in row:
            cell.fill = fill

    fill = PatternFill(start_color="FF00FF", end_color="FF00FF", fill_type="solid")
    for row in worksheet.iter_rows(min_row=1, max_row=1, min_col=1 + stopAsc, max_col=stopTrans):
        for cell in row:
            cell.fill = fill

    fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    for row in worksheet.iter_rows(min_row=1, max_row=1, min_col=1 + stopTrans, max_col=stopDesc):
        for cell in row:
            cell.fill = fill

    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for row in worksheet.iter_rows(min_row=1, max_row=1, min_col=1 + stopDesc, max_col=stopSig):
        for cell in row:
            cell.fill = fill

    fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
    for row in worksheet.iter_rows(min_row=1, max_row=1, min_col=1 + stopSig, max_col=stopRect):
        for cell in row:
            cell.fill = fill

    # colour all the comments
    fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    # Iterate through the rows, starting from the second row
    rownumber = 0
    for row in worksheet.iter_rows(min_row=0, values_only=True):
        rownumber += 1
        if row[2] is None and rownumber > 1:
            for row in worksheet.iter_rows(min_row=rownumber, max_row=rownumber, min_col=0, max_col=2):
                for cell in row:
                    cell.fill = fill

    xlsx_file = name + "_detected" + ".xlsx"
    workbook.save(xlsx_file)


def exportToCsv(data, filename):
    with open(filename + '_detected' + '.csv', 'w', newline='') as csv_file:
        csv_writer = csv.writer(csv_file)
        for row in data:
            csv_writer.writerow(row)